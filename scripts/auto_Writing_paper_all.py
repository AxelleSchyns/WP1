from argparse import ArgumentParser
import re
import numpy as np
from openpyxl import load_workbook, Workbook
import os

def get_unique_sheet_name(wb, base_name):
    """Generate a unique sheet name if base_name already exists."""
    existing_sheets = wb.sheetnames
    if base_name not in existing_sheets:
        return base_name  # Name is unique, use it

    # Find the next available numeric identifier
    count = 1
    while f"{base_name}_{count}" in existing_sheets:
        count += 1
    return f"{base_name}_{count}"

# Allows to write in an excel file all results from either uliege, cam or crc file for all models at once
# Writes in a new sheet of an existing excel file or creates a new one if it does not exist

# ---------- Arguments ----------
parser = ArgumentParser()
parser.add_argument(
    "--log_file",
    type=str,
    required=True,
    help="Path to the log file",
)
parser.add_argument(
    "--excel_file",
    type=str,
    required=True,
    help="Path to the Excel file",
)
parser.add_argument(
    "--sheet_name",
    type=str,
    required=False,
    help="Name of the sheet to write in",
)
parser.add_argument(
    "--nb_models",
    type=int,
    required=False,
    help="Number of models to write in the excel file",
)
parser.add_argument(
    "--dataset",
    type=str,
    required=False,
    default="Uliege - all",
    help="Name of the dataset",
)
args = parser.parse_args()


# ---------- Setting up of Excel file ----------
# Check if the file exists
if os.path.exists(args.excel_file):
    wb = load_workbook(args.excel_file)  # Load existing file
    print(f"File '{args.excel_file}' found. Adding a new sheet.")
else:
    wb = Workbook()  # Create a new workbook
    print(f"File '{args.excel_file}' not found. Creating a new Excel file.")

# Ensure sheet name is unique
unique_sheet_name = get_unique_sheet_name(wb, args.sheet_name)
ws = wb.create_sheet(unique_sheet_name)  # Create the new sheet
print(f"Sheet '{unique_sheet_name}' added successfully.")

# Merge cells
ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=6)   # Slot 1 (A1:E1)
ws["A2"] = args.dataset
ws.cell(row=2, column=1, value="Model")
ws.cell(row = 2, column = 2, value= "T_indexing")
ws.cell(row = 2, column = 3, value= "T_search")
ws.cell(row = 2, column = 4, value= "Top-1")
ws.cell(row = 2, column = 5, value= "Top-5")
ws.cell(row = 2, column = 6, value= "Maj")

# ---------- Extracting from log file ----------
log_file = args.log_file

# Read the log file
with open(log_file, "r") as file:
    log_content = file.read()

# ---------- Names -------------------------------
# Extract all model names using regular expression
pattern = r'--\s*([^-].*?)\s*--'

matches = re.findall(pattern, log_content)
filtered_matches = [match for match in matches if match.strip()]
if len(filtered_matches) != args.nb_models:
    print(f"Expected {args.nb_models} models, but found {len(filtered_matches)} in the log file.")
    exit(1)

# Write model names in the excel file
for i in range(args.nb_models):
    ws.cell(row = 3 + i, column = 1, value= filtered_matches[i])

# ---------- Numbers -----------------------------------
# Extract all numbers using regular expression
numbers = re.findall(r'\b(?:0(?:\.\d+)?|\d+\.\d+|\d+e[+-]?\d+)\b', log_content) #re.findall(r'\b\d+\.\d+\b', log_content)

print(str(len(numbers)) + " numbers were retrieved")
mul = int(len(numbers) / args.nb_models)

for j in range(args.nb_models):
    ws.cell(row = 3 + j, column = 2, value= np.round(np.float64(numbers[j * mul +3]), 2))
    ws.cell(row = 3 + j, column = 3, value= np.round(np.float64(numbers[j * mul + 17]), 2))
    ws.cell(row = 3 + j, column = 4, value= np.round(np.float64(numbers[j * mul + 4])*100, 2))
    ws.cell(row = 3 + j, column = 5, value= np.round(np.float64(numbers[j * mul + 5])*100, 2))
    ws.cell(row = 3 + j, column = 6, value= np.round(np.float64(numbers[j * mul + 10])*100, 2))

wb.save(args.excel_file)
